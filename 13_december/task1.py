# Записать все 3 файла в один новый общий на одной страницу
# Прочитать 3 файла -> записываем в массивы. Записываем массивы в один файл

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any


def parse_workbook(workbook: Workbook) -> list[list[Any]]:
    worksheet = workbook.active
    data = []

    for row in worksheet:
        data.append([cell.value for cell in row])

    return data


def  write_data_to_worksheet(worksheet: Worksheet, data: list[list[Any]]) -> None:
    for j, row in enumerate(data, start = 1):
        for i, value in enumerate(row, start = 1):
            worksheet.cell(row = j, column = i, value = value)


if __name__ == '__main__':
    workbook1 = load_workbook(filename='13_december/Книга1.xlsx')
    workbook2 = load_workbook(filename='13_december/Книга2.xlsx')
    workbook3 = load_workbook(filename='13_december/Книга3.xlsx')

    workbook = Workbook()
    
    write_data_to_worksheet(workbook.active, parse_workbook(workbook1))
    write_data_to_worksheet(workbook.active, parse_workbook(workbook2))
    write_data_to_worksheet(workbook.active, parse_workbook(workbook3))

    workbook.save('13_december/Result.xlsx')